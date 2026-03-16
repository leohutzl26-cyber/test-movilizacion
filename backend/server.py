import google.generativeai as genai
from fastapi import FastAPI, APIRouter, HTTPException, Depends, UploadFile, File
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
import asyncio
import uuid
import base64
import string
import random
from pathlib import Path
from pydantic import BaseModel, Field, EmailStr
from typing import List, Optional, Any
from datetime import datetime, timezone, timedelta
from passlib.context import CryptContext
from jose import jwt, JWTError
import resend
from fastapi.responses import Response, StreamingResponse
import io

# Report dependencies
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from reportlab.lib.pagesizes import landscape, letter
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
except ImportError as e:
    logging.warning(f"Librerías de reportes no instaladas: {e}")

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# JWT Config
SECRET_KEY = os.environ.get('JWT_SECRET', str(uuid.uuid4()))
ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE = 24

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")

resend_api_key = os.environ.get('RESEND_API_KEY')
if resend_api_key:
    resend.api_key = resend_api_key
SENDER_EMAIL = os.environ.get('SENDER_EMAIL', 'onboarding@resend.dev')

app = FastAPI()

# ============ CORS CONFIGURATION (MUST BE TOP LEVEL) ============
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # Permitimos todos para evitar problemas de matching en Render
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
    expose_headers=["Content-Disposition"], # Importante para descargas
)

@app.get("/")
async def root():
    return {"status": "ok", "message": "Movilizacion HCU Backend Running"}

api_router = APIRouter(prefix="/api")
security = HTTPBearer()

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# ============ PYDANTIC MODELS ============

class UserRegister(BaseModel):
    email: EmailStr
    password: str
    name: str
    role: str = "solicitante"

class UserLogin(BaseModel):
    email: EmailStr
    password: str

class ForgotPassword(BaseModel):
    email: EmailStr

class ResetPassword(BaseModel):
    token: str
    new_password: str

class ChangePassword(BaseModel):
    current_password: str
    new_password: str

class VehicleCreate(BaseModel):
    plate: str
    brand: str
    model: str
    type: str = "Auto/SUV"
    year: int = 2024
    mileage: float = 0
    next_maintenance_km: float = 10000

class VehicleStatusUpdate(BaseModel):
    status: str

class VehicleMileageUpdate(BaseModel):
    mileage: float

class TripCreate(BaseModel):
    origin: str
    destination: str
    patient_name: str = ""
    patient_unit: str = ""
    priority: str = "normal"
    notes: str = ""
    trip_type: str = "no_clinico"
    clinical_team: str = ""
    contact_person: str = ""
    scheduled_date: str = ""
    rut: str = ""
    age: str = ""
    diagnosis: str = ""
    weight: str = ""
    bed: str = ""
    transfer_reason: str = ""
    requester_person: str = ""
    attending_physician: str = ""
    appointment_time: str = ""
    departure_time: str = ""
    required_personnel: List[str] = Field(default_factory=list)
    patient_requirements: List[str] = Field(default_factory=list)
    accompaniment: str = ""
    task_details: str = ""
    staff_count: str = ""
    accompaniment_staff_id: Optional[str] = None
    assigned_clinical_staff: List[dict] = Field(default_factory=list)  # [{type, staff_id, staff_name}]
    
class ClinicalStaffCreate(BaseModel):
    name: str
    role: str
    is_active: bool = True

class ClinicalStaffUpdate(BaseModel):
    name: Optional[str] = None
    role: Optional[str] = None
    is_active: Optional[bool] = None

class TripStatusUpdate(BaseModel):
    status: str
    mileage: Optional[float] = None
    vehicle_id: Optional[str] = None
    cancel_reason: Optional[str] = None
    
class TripGroupUpdate(BaseModel):
    group_id: str
    order_in_group: int = 0
    
class TripReorder(BaseModel):
    trip_ids: List[str]
    
class TripUpdate(BaseModel):
    origin: Optional[str] = None
    destination: Optional[str] = None
    patient_name: Optional[str] = None
    patient_unit: Optional[str] = None
    priority: Optional[str] = None
    notes: Optional[str] = None
    trip_type: Optional[str] = None
    scheduled_date: Optional[str] = None
    clinical_team: Optional[str] = None
    rut: Optional[str] = None
    age: Optional[str] = None
    diagnosis: Optional[str] = None
    weight: Optional[str] = None
    bed: Optional[str] = None
    transfer_reason: Optional[str] = None
    requester_person: Optional[str] = None
    attending_physician: Optional[str] = None
    appointment_time: Optional[str] = None
    departure_time: Optional[str] = None
    required_personnel: Optional[List[str]] = None
    patient_requirements: Optional[List[str]] = None
    accompaniment: Optional[str] = None
    task_details: Optional[str] = None
    staff_count: Optional[str] = None
    accompaniment_staff_id: Optional[str] = None
    assigned_clinical_staff: Optional[List[dict]] = None

class ManagerAssign(BaseModel):
    driver_id: str
    vehicle_id: Optional[str] = None

class DestinationCreate(BaseModel):
    name: str
    address: str = ""

class DestinationUpdate(BaseModel):
    name: Optional[str] = None
    address: Optional[str] = None
    is_active: Optional[bool] = None

class ClinicalTeamUpdate(BaseModel):
    clinical_team: str

class OriginServiceCreate(BaseModel):
    name: str
    is_active: bool = True

class OriginServiceUpdate(BaseModel):
    name: Optional[str] = None
    is_active: Optional[bool] = None

class UserRoleUpdate(BaseModel):
    role: str

class DriverLicenseUpdate(BaseModel):
    license_expiry: str

# --- MODELOS BITÁCORA ---
class VehicleChecklist(BaseModel):
    vehicle_id: str
    fuel_level: str
    lights_ok: bool
    tires_ok: bool
    siren_ok: bool
    clean_interior: bool
    observations: str = ""

class FuelRecord(BaseModel):
    vehicle_id: str
    mileage: float
    liters: float
    amount: float
    receipt_number: str = ""

class IncidentRecord(BaseModel):
    vehicle_id: str
    incident_type: str
    description: str
    severity: str

# ============ AUTH UTILITIES ============

def create_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    expire = datetime.now(timezone.utc) + (expires_delta or timedelta(hours=ACCESS_TOKEN_EXPIRE))
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)

async def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)):
    try:
        payload = jwt.decode(credentials.credentials, SECRET_KEY, algorithms=[ALGORITHM])
        user_id = payload.get("sub")
        if not user_id:
            raise HTTPException(status_code=401, detail="Token invalido")
        user = await db.users.find_one({"id": user_id}, {"_id": 0})
        if not user:
            raise HTTPException(status_code=401, detail="Usuario no encontrado")
        if user.get("status") != "aprobado":
            raise HTTPException(status_code=403, detail="Usuario pendiente de aprobacion")
        user["license_expired"] = False
        if user.get("role") == "conductor" and user.get("license_expiry"):
            try:
                expiry = datetime.fromisoformat(user["license_expiry"])
                if expiry.tzinfo is None:
                    expiry = expiry.replace(tzinfo=timezone.utc)
                if expiry < datetime.now(timezone.utc):
                    user["license_expired"] = True
            except (ValueError, TypeError):
                pass
        return user
    except JWTError:
        raise HTTPException(status_code=401, detail="Token invalido o expirado")

def require_roles(*roles):
    async def role_checker(user=Depends(get_current_user)):
        if user["role"] not in roles:
            raise HTTPException(status_code=403, detail="No tiene permisos para esta accion")
        return user
    return role_checker

async def log_action(user_id: str, user_name: str, user_role: str, action: str, entity_type: str, entity_id: str = "", details: str = ""):
    entry = {
        "id": str(uuid.uuid4()),
        "user_id": user_id,
        "user_name": user_name,
        "user_role": user_role,
        "action": action,
        "entity_type": entity_type,
        "entity_id": entity_id,
        "details": details,
        "timestamp": datetime.now(timezone.utc).isoformat()
    }
    print(f"AUDIT LOG: {action} on {entity_type} ({entity_id}) by {user_name} - {details}")
    try:
        await db.audit_logs.insert_one(entry)
    except Exception as e:
        logger.error(f"Error saving audit log: {e}")

# ============ ENDPOINTS GENERALES ============

@api_router.post("/auth/register")
async def register(data: UserRegister):
    existing = await db.users.find_one({"email": data.email}, {"_id": 0})
    if existing: raise HTTPException(status_code=400, detail="El correo ya esta registrado")
    user = {
        "id": str(uuid.uuid4()),
        "email": data.email,
        "password_hash": pwd_context.hash(data.password),
        "name": data.name,
        "role": data.role,
        "status": "pendiente",
        "shift_type": None,
        "extra_available": False,
        "license_expiry": None,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.users.insert_one(user)
    await log_action(str(user["id"]), data.name, data.role, "registro", "usuario", str(user["id"]), f"Nuevo usuario registrado: {data.email}")
    return {"message": "Registro exitoso. Pendiente de aprobacion.", "user_id": user["id"]}

@api_router.post("/auth/login")
async def login(data: UserLogin):
    user = await db.users.find_one({"email": data.email}, {"_id": 0})
    if not user or not pwd_context.verify(data.password, user["password_hash"]):
        raise HTTPException(status_code=401, detail="Credenciales incorrectas")
    if user["status"] != "aprobado":
        raise HTTPException(status_code=403, detail="Cuenta pendiente de aprobacion")
        
    token = create_token({"sub": user["id"], "role": user["role"]})
    return {
        "token": token,
        "user": {
            "id": user["id"], "email": user["email"], "name": user["name"],
            "role": user["role"], "shift_type": user.get("shift_type"), "extra_available": user.get("extra_available", False)
        }
    }

@api_router.post("/auth/forgot-password")
async def forgot_password(data: ForgotPassword):
    user = await db.users.find_one({"email": data.email}, {"_id": 0})
    if not user: return {"message": "Instrucciones enviadas."}
    reset_code = ''.join(random.choices(string.digits, k=6))
    await db.users.update_one({"id": user["id"]}, {"$set": {"reset_token": reset_code}})
    if resend_api_key:
        try:
            html_content = f"<h2>Recuperacion de Contrasena</h2><p>Codigo: <b>{reset_code}</b></p>"
            params = {"from": SENDER_EMAIL, "to": [data.email], "subject": "Recuperacion de Contrasena", "html": html_content}
            await asyncio.to_thread(resend.Emails.send, params)
        except Exception as e: pass
    return {"message": "Instrucciones enviadas", "reset_token": reset_code}

@api_router.post("/auth/reset-password")
async def reset_password(data: ResetPassword):
    user = await db.users.find_one({"reset_token": data.token})
    if not user: raise HTTPException(status_code=400, detail="Código inválido")
    new_hash = pwd_context.hash(data.new_password)
    await db.users.update_one({"id": user["id"]}, {"$set": {"password_hash": new_hash, "reset_token": None}})
    return {"message": "Contrasena actualizada"}

@api_router.put("/auth/change-password")
async def change_password(data: ChangePassword, user=Depends(get_current_user)):
    db_user = await db.users.find_one({"id": user["id"]}, {"_id": 0})
    if not pwd_context.verify(data.current_password, db_user.get("password_hash", "")):
        raise HTTPException(status_code=400, detail="La contraseña actual es incorrecta")
    new_hash = pwd_context.hash(data.new_password)
    await db.users.update_one({"id": user["id"]}, {"$set": {"password_hash": new_hash}})
    return {"message": "Contraseña actualizada exitosamente"}

@api_router.get("/auth/me")
@api_router.get("/me")
async def get_me(user=Depends(get_current_user)):
    return {k: v for k, v in user.items() if k != "password_hash"}

@api_router.get("/users")
async def list_users(user=Depends(require_roles("admin"))):
    return await db.users.find({}, {"_id": 0, "password_hash": 0}).to_list(1000)

@api_router.put("/users/{user_id}/approve")
async def approve_user(user_id: str, user=Depends(require_roles("admin"))):
    target_user = await db.users.find_one({"id": user_id})
    if target_user:
        await db.users.update_one({"id": user_id}, {"$set": {"status": "aprobado"}})
        await log_action(user["id"], user["name"], user["role"], "aprobar_usuario", "usuario", user_id, f"Aprobado: {target_user.get('email')}")
    return {"message": "Aprobado"}

@api_router.put("/users/{user_id}/reject")
async def reject_user(user_id: str, user=Depends(require_roles("admin"))):
    target_user = await db.users.find_one({"id": user_id})
    if target_user:
        await db.users.update_one({"id": user_id}, {"$set": {"status": "rechazado"}})
        await log_action(user["id"], user["name"], user["role"], "rechazar_usuario", "usuario", user_id, f"Rechazado: {target_user.get('email')}")
    return {"message": "Rechazado"}

@api_router.put("/users/{user_id}/role")
async def update_role(user_id: str, data: UserRoleUpdate, user=Depends(require_roles("admin"))):
    target_user = await db.users.find_one({"id": user_id})
    if target_user:
        await db.users.update_one({"id": user_id}, {"$set": {"role": data.role}})
        await log_action(user["id"], user["name"], user["role"], "cambiar_rol", "usuario", user_id, f"Rol de {target_user.get('email')} cambiado a {data.role}")
    return {"message": "Rol actualizado"}

@api_router.delete("/users/{user_id}")
async def delete_user(user_id: str, user=Depends(require_roles("admin"))):
    await db.users.delete_one({"id": user_id})
    return {"message": "Eliminado"}

@api_router.get("/drivers")
async def list_drivers(user=Depends(require_roles("admin", "coordinador", "gestion_camas"))):
    drivers = await db.users.find({"role": "conductor"}, {"_id": 0, "password_hash": 0}).to_list(1000)
    vehicles = await db.vehicles.find({}, {"_id": 0}).to_list(1000)
    v_map = {v["id"]: v for v in vehicles}
    for d in drivers:
        if d.get("vehicle_id"):
            v = v_map.get(d["vehicle_id"])
            if v:
                d["vehicle_plate"] = v.get("plate", "")
                d["vehicle_type"] = v.get("type", "Auto/SUV")
    return drivers

@api_router.put("/drivers/{driver_id}/extra-availability")
async def toggle_extra_availability(driver_id: str, user=Depends(get_current_user)):
    driver = await db.users.find_one({"id": driver_id})
    new_val = not driver.get("extra_available", False)
    await db.users.update_one({"id": driver_id}, {"$set": {"extra_available": new_val}})
    return {"message": "Ok", "extra_available": new_val}

@api_router.put("/drivers/{driver_id}/license")
async def update_license(driver_id: str, data: DriverLicenseUpdate, user=Depends(require_roles("admin"))):
    await db.users.update_one({"id": driver_id, "role": "conductor"}, {"$set": {"license_expiry": data.license_expiry}})
    return {"message": "Ok"}

@api_router.get("/vehicles")
async def list_vehicles(user=Depends(get_current_user)):
    vehicles = await db.vehicles.find({}, {"_id": 0}).to_list(1000)
    for v in vehicles:
        if v.get("status") == "en_uso":
            # Buscar el viaje activo para este vehículo
            active_trip = await db.trips.find_one({
                "vehicle_id": v["id"],
                "status": "en_curso"
            }, {"_id": 0, "driver_name": 1, "destination": 1, "clinical_team": 1})
            if active_trip:
                v["current_driver"] = active_trip.get("driver_name")
                v["current_destination"] = active_trip.get("destination")
                v["current_clinical_team"] = active_trip.get("clinical_team")
    return vehicles

@api_router.post("/vehicles")
async def create_vehicle(data: VehicleCreate, user=Depends(require_roles("admin"))):
    vehicle = data.model_dump()
    vehicle["id"] = str(uuid.uuid4())
    vehicle["status"] = "disponible"
    vehicle["maintenance_alert"] = None
    vehicle["created_at"] = datetime.now(timezone.utc).isoformat()
    await db.vehicles.insert_one(vehicle)
    vehicle.pop("_id", None)
    return vehicle

@api_router.put("/vehicles/{vehicle_id}")
async def update_vehicle(vehicle_id: str, data: VehicleCreate, user=Depends(require_roles("admin"))):
    await db.vehicles.update_one({"id": vehicle_id}, {"$set": data.model_dump()})
    return {"message": "Ok"}

@api_router.put("/vehicles/{vehicle_id}/status")
async def update_vehicle_status(vehicle_id: str, data: VehicleStatusUpdate, user=Depends(require_roles("admin", "coordinador", "conductor"))):
    await db.vehicles.update_one({"id": vehicle_id}, {"$set": {"status": data.status}})
    return {"message": "Ok"}

@api_router.put("/vehicles/{vehicle_id}/mileage")
async def update_vehicle_mileage(vehicle_id: str, data: VehicleMileageUpdate, user=Depends(require_roles("admin", "conductor"))):
    vehicle = await db.vehicles.find_one({"id": vehicle_id})
    next_maint = vehicle.get("next_maintenance_km", 10000)
    diff = next_maint - data.mileage
    alert = "rojo" if diff <= 0 else "amarillo" if diff <= 1000 else None
    await db.vehicles.update_one({"id": vehicle_id}, {"$set": {"mileage": data.mileage, "maintenance_alert": alert}})
    return {"message": "Ok"}

@api_router.delete("/vehicles/{vehicle_id}")
async def delete_vehicle(vehicle_id: str, user=Depends(require_roles("admin"))):
    await db.vehicles.delete_one({"id": vehicle_id})
    return {"message": "Ok"}

# ============ TRIP MANAGEMENT ============

@api_router.post("/trips")
async def create_trip(data: TripCreate, user=Depends(require_roles("solicitante", "coordinador", "admin"))):
    today_str = datetime.now(timezone.utc).strftime("%y%m%d")
    today_prefix = f"TR-{today_str}-"
    count_today = await db.trips.count_documents({"tracking_number": {"$regex": f"^{today_prefix}"}})
    sequential = count_today + 1
    folio = f"{today_prefix}{sequential:03d}"
    while await db.trips.find_one({"tracking_number": folio}):
        sequential += 1
        folio = f"{today_prefix}{sequential:03d}"

    initial_status = "pendiente"
    if data.trip_type == "clinico" and user["role"] == "solicitante":
        initial_status = "revision_gestor"

    trip = {
        "id": str(uuid.uuid4()), "tracking_number": folio, "requester_id": user["id"], "requester_name": user["name"],
        "driver_id": None, "driver_name": None, "origin": data.origin, "destination": data.destination,
        "patient_name": data.patient_name, "patient_unit": data.patient_unit, "priority": data.priority,
        "status": initial_status, "group_id": None, "order_in_group": 0, "notes": data.notes, "trip_type": data.trip_type,
        "scheduled_date": data.scheduled_date or datetime.now(timezone.utc).strftime("%Y-%m-%d"),
        "vehicle_id": None, "start_mileage": None, "end_mileage": None,
        "created_at": datetime.now(timezone.utc).isoformat(), "updated_at": datetime.now(timezone.utc).isoformat(),
        "completed_at": None, "rut": data.rut, "age": data.age, "diagnosis": data.diagnosis, "weight": data.weight,
        "bed": data.bed, "transfer_reason": data.transfer_reason, "requester_person": user["name"],
        "attending_physician": data.attending_physician, "appointment_time": data.appointment_time, "departure_time": data.departure_time,
        "required_personnel": data.required_personnel, "patient_requirements": data.patient_requirements,
        "accompaniment": data.accompaniment, "accompaniment_staff_id": data.accompaniment_staff_id,
        "task_details": data.task_details, "staff_count": data.staff_count,
        "clinical_team": data.clinical_team,
        "assigned_clinical_staff": data.assigned_clinical_staff,
    }
    await db.trips.insert_one(trip)
    trip.pop("_id", None)
    await log_action(user["id"], user["name"], user["role"], "crear_traslado", "traslado", trip["id"], f"Folio {folio}: {data.origin} -> {data.destination}")
    return trip

@api_router.get("/trips")
async def list_trips(user=Depends(get_current_user)):
    try:
        query = {}
        if user["role"] == "solicitante": query["requester_id"] = user["id"]
        elif user["role"] == "conductor": query["driver_id"] = user["id"]
        trips = await db.trips.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
        return trips
    except Exception as e:
        logger.error(f"Error list_trips: {e}")
        return []

@api_router.get("/trips/pool")
async def trip_pool(user=Depends(require_roles("conductor", "coordinador", "gestion_camas"))):
    return await db.trips.find(
        {"status": "pendiente", "$or": [{"driver_id": None}, {"driver_id": {"$exists": False}}]},
        {"_id": 0}
    ).sort("created_at", -1).to_list(1000)

@api_router.get("/trips/gestion_revision")
async def trips_for_gestion_revision(user=Depends(require_roles("gestion_camas", "admin"))):
    return await db.trips.find({"status": "revision_gestor"}, {"_id": 0}).sort("created_at", -1).to_list(1000)

@api_router.put("/trips/{trip_id}/approve-gestor")
async def approve_trip_gestor(trip_id: str, data: TripUpdate, user=Depends(require_roles("gestion_camas", "admin"))):
    trip = await db.trips.find_one({"id": trip_id})
    if not trip: raise HTTPException(status_code=404, detail="Viaje no encontrado")
    if trip.get("status") != "revision_gestor":
        raise HTTPException(status_code=400, detail="El viaje no está en revisión")
        
    update_data = {k: v for k, v in data.model_dump().items() if v is not None}
    update_data["status"] = "pendiente"
    update_data["updated_at"] = datetime.now(timezone.utc).isoformat()

    # Si se actualizó la lista detallada, recalculamos el string de clinical_team
    if "assigned_clinical_staff" in update_data and update_data["assigned_clinical_staff"]:
        staff_names = [f"{s.get('type')}: {s.get('staff_name') or 'Por identificar'}" for s in update_data["assigned_clinical_staff"]]
        update_data["clinical_team"] = ", ".join(staff_names)

    await db.trips.update_one({"id": trip_id}, {"$set": update_data})
    
    # Send a notification action to the audit logs
    await log_action(user["id"], user["name"], user["role"], "aprobar", "traslado", trip["id"], f"Folio {trip['tracking_number']} aprobado por gestor")
    
    return {"message": "Viaje aprobado correctamente"}

@api_router.get("/trips/active")
async def active_trips(user=Depends(require_roles("coordinador", "admin", "gestion_camas"))):
    try:
        trips = await db.trips.find({"status": {"$in": ["pendiente", "asignado", "en_curso"]}}, {"_id": 0}).sort("created_at", -1).to_list(1000)
        for t in trips:
            if t.get("vehicle_id"):
                try:
                    v = await db.vehicles.find_one({"id": t["vehicle_id"]}, {"_id": 0, "plate": 1, "type": 1})
                    if v:
                        t["vehicle_plate"] = v.get("plate", "")
                        t["vehicle_type"] = v.get("type", "Auto/SUV")
                except: pass
        return trips
    except Exception as e:
        logger.error(f"Error active_trips: {e}")
        return []

@api_router.get("/trips/history")
async def trips_history(
    status: Optional[str] = None,
    trip_type: Optional[str] = None,
    patient_name: Optional[str] = None,
    folio: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    user=Depends(require_roles("coordinador", "admin", "gestion_camas"))
):
    try:
        query: dict[str, Any] = {}
        if status: query["status"] = status
        if trip_type: query["trip_type"] = trip_type
        if patient_name: query["patient_name"] = {"$regex": patient_name, "$options": "i"}
        if folio: query["tracking_number"] = {"$regex": folio, "$options": "i"}
        
        if start_date or end_date:
            query["scheduled_date"] = {}
            if start_date: query["scheduled_date"]["$gte"] = start_date
            if end_date: query["scheduled_date"]["$lte"] = end_date

        trips = await db.trips.find(query, {"_id": 0}).sort("created_at", -1).to_list(5000)
        vehicle_ids = list(set(t.get("vehicle_id") for t in trips if t.get("vehicle_id")))
        
        vehicles_map = {}
        if vehicle_ids:
            try:
                found_vehs = await db.vehicles.find({"id": {"$in": vehicle_ids}}, {"_id": 0, "id": 1, "plate": 1, "type": 1}).to_list(len(vehicle_ids))
                for v in found_vehs:
                    if "id" in v:
                        vehicles_map[v["id"]] = {"plate": v.get("plate", ""), "type": v.get("type", "Auto/SUV")}
            except Exception as ev:
                logger.error(f"Error lookup vehicles map: {ev}")

        for t in trips: 
            v_info = vehicles_map.get(t.get("vehicle_id"), {})
            t["vehicle_plate"] = v_info.get("plate", "")
            t["vehicle_type"] = v_info.get("type", "Auto/SUV")
        return trips
    except Exception as e:
        logger.error(f"Error trips_history: {e}")
        traceback.print_exc()
        return []

# --- AQUI AGREGAMOS PERMISO PARA EL CALENDARIO ---
@api_router.get("/trips/calendar")
async def trips_calendar(start_date: str = None, end_date: str = None, user=Depends(require_roles("coordinador", "admin", "gestion_camas"))):
    query = {"status": {"$ne": "cancelado"}}
    if start_date and end_date: query["scheduled_date"] = {"$gte": start_date, "$lte": end_date}
    return await db.trips.find(query, {"_id": 0}).sort("scheduled_date", 1).to_list(1000)

@api_router.put("/trips/reorder")
async def reorder_trips(data: TripReorder, user=Depends(require_roles("coordinador", "admin", "gestion_camas"))):
    for i, t_id in enumerate(data.trip_ids):
        await db.trips.update_one({"id": t_id}, {"$set": {"order_in_group": i}})
    await log_action(user["id"], user["name"], user["role"], "reordenar", "traslados", "multi", f"Reordenados {len(data.trip_ids)} traslados")
    return {"message": "Ok"}

@api_router.get("/trips/by-vehicle")
async def trips_by_vehicle(date: str = None, user=Depends(require_roles("coordinador", "admin", "gestion_camas"))):
    target_date = date or datetime.now(timezone.utc).strftime("%Y-%m-%d")
    vehicles = await db.vehicles.find({}, {"_id": 0}).to_list(500)
    trips = await db.trips.find({"scheduled_date": target_date, "status": {"$ne": "cancelado"}}, {"_id": 0}).sort("order_in_group", 1).to_list(5000)
    res = [{"vehicle": v, "trips": [t for t in trips if t.get("vehicle_id") == v["id"]]} for v in vehicles]
    unassigned = [t for t in trips if not t.get("vehicle_id")]
    if unassigned: res.append({"vehicle": {"id": "unassigned", "plate": "Sin Vehiculo"}, "trips": unassigned})
    return res

@api_router.get("/trips/by-driver")
async def trips_by_driver(date: str = None, user=Depends(require_roles("coordinador", "admin", "gestion_camas"))):
    target_date = date or datetime.now(timezone.utc).strftime("%Y-%m-%d")
    # Obtener conductores y sus veículos (si tienen en su perfil)
    drivers = await db.users.find({"role": "conductor", "status": "aprobado"}, {"id": 1, "name": 1, "vehicle_id": 1, "_id": 0}).to_list(500)
    vehicles = await db.vehicles.find({}, {"_id": 0}).to_list(500)
    v_map = {v["id"]: v for v in vehicles}
    
    trips = await db.trips.find({"scheduled_date": target_date, "status": {"$ne": "cancelado"}}, {"_id": 0}).sort("order_in_group", 1).to_list(5000)
    
    res = []
    for d in drivers:
        d_veh = v_map.get(d.get("vehicle_id"))
        if d_veh:
            d["vehicle_plate"] = d_veh["plate"]
            d["vehicle_type"] = d_veh.get("type", "Auto/SUV")
        
        driver_trips = [t for t in trips if t.get("driver_id") == d["id"]]
        # Asegurar que cada viaje también tenga el tipo de veículo asignado
        for t in driver_trips:
            t_veh = v_map.get(t.get("vehicle_id"))
            if t_veh:
                t["vehicle_type"] = t_veh.get("type", "Auto/SUV")
                
        res.append({"driver": d, "trips": driver_trips})
        
    unassigned = [t for t in trips if not t.get("driver_id")]
    for t in unassigned:
        t_veh = v_map.get(t.get("vehicle_id"))
        if t_veh:
            t["vehicle_type"] = t_veh.get("type", "Auto/SUV")
            
    if unassigned: 
        res.append({"driver": {"id": "unassigned", "name": "Sin Conductor"}, "trips": unassigned})
        
    return res


@api_router.get("/trips/{trip_id}")
async def get_trip_detail(trip_id: str, user=Depends(get_current_user)):
    return await db.trips.find_one({"id": trip_id}, {"_id": 0})

@api_router.get("/trips/{trip_id}/audit")
async def get_trip_audit(trip_id: str, user=Depends(require_roles("admin", "coordinador"))):
    return await db.audit_logs.find({"entity_id": trip_id}, {"_id": 0}).sort("timestamp", 1).to_list(1000)

@api_router.put("/trips/{trip_id}")
async def edit_trip(trip_id: str, data: TripUpdate, user=Depends(get_current_user)):
    update_data = {k: v for k, v in data.model_dump().items() if v is not None}
    update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
    
    # Sincronizar clinical_team si se envía personal asignado
    if "assigned_clinical_staff" in update_data and update_data["assigned_clinical_staff"]:
        staff_names = [f"{s.get('type')}: {s.get('staff_name') or 'Por identificar'}" for s in update_data["assigned_clinical_staff"]]
        update_data["clinical_team"] = ", ".join(staff_names)
        
    await db.trips.update_one({"id": trip_id}, {"$set": update_data})
    
    trip = await db.trips.find_one({"id": trip_id})
    folio = trip.get("tracking_number", "N/A")
    fields = ", ".join(update_data.keys())
    await log_action(user["id"], user["name"], user["role"], "editar", "traslado", trip_id, f"Actualizados campos ({fields}) en folio {folio}")
    
    return {"message": "Viaje actualizado"}

@api_router.put("/trips/{trip_id}/manager-assign")
async def manager_assign_trip(trip_id: str, data: ManagerAssign, user=Depends(require_roles("coordinador", "admin", "gestion_camas"))):
    trip = await db.trips.find_one({"id": trip_id})
    driver = await db.users.find_one({"id": data.driver_id, "role": "conductor"}, {"_id": 0})
    update_data = {"driver_id": data.driver_id, "driver_name": driver["name"], "status": "asignado"}
    if data.vehicle_id: update_data["vehicle_id"] = data.vehicle_id
    await db.trips.update_one({"id": trip_id}, {"$set": update_data})
    
    await log_action(user["id"], user["name"], user["role"], "manager_assign", "traslado", trip_id, f"Folio {trip['tracking_number']} asignado a conductor {driver['name']}")
    return {"message": "Asignado"}

@api_router.put("/trips/{trip_id}/unassign")
async def unassign_trip(trip_id: str, user=Depends(require_roles("coordinador", "admin", "conductor", "gestion_camas"))):
    trip = await db.trips.find_one({"id": trip_id}, {"_id": 0})
    update_data = {"status": "pendiente", "driver_id": None, "driver_name": None, "vehicle_id": None}
    if trip and trip.get("driver_id") and user["role"] == "conductor":
        update_data["previous_driver_id"] = trip["driver_id"]
        update_data["previous_driver_name"] = trip.get("driver_name")
        update_data["returned_at"] = datetime.now(timezone.utc).isoformat()
    await db.trips.update_one({"id": trip_id}, {"$set": update_data})
    
    await log_action(user["id"], user["name"], user["role"], "desasignar", "traslado", trip_id, f"Folio {trip['tracking_number']} desasignado (Sin conductor)")
    return {"message": "Desasignado"}
    
@api_router.put("/trips/{trip_id}/assign")
async def assign_trip(trip_id: str, user=Depends(require_roles("conductor"))):
    trip = await db.trips.find_one({"id": trip_id})
    await db.trips.update_one({"id": trip_id}, {"$set": {"driver_id": user["id"], "driver_name": user["name"], "status": "asignado"}})
    await log_action(user["id"], user["name"], user["role"], "tomar_traslado", "traslado", trip_id, f"Folio {trip['tracking_number']} tomado por conductor {user['name']}")
    return {"message": "Asignado"}

@api_router.put("/trips/{trip_id}/clinical-team")
async def assign_clinical_team(trip_id: str, data: ClinicalTeamUpdate, user=Depends(require_roles("admin", "coordinador", "gestion_camas", "solicitante"))):
    trip = await db.trips.find_one({"id": trip_id})
    await db.trips.update_one({"id": trip_id}, {"$set": {"clinical_team": data.clinical_team, "updated_at": datetime.now(timezone.utc).isoformat()}})
    await log_action(user["id"], user["name"], user["role"], "equipo_clinico", "traslado", trip_id, f"Asignado equipo clínico: {data.clinical_team} en folio {trip['tracking_number']}")
    return {"message": "Personal clínico asignado correctamente"}

@api_router.put("/trips/{trip_id}/status")
async def update_trip_status(trip_id: str, data: TripStatusUpdate, user=Depends(get_current_user)):
    trip = await db.trips.find_one({"id": trip_id})
    if not trip: raise HTTPException(status_code=404, detail="Viaje no encontrado")
    update_data = {"status": data.status, "updated_at": datetime.now(timezone.utc).isoformat()}
    if data.status == "cancelado" and data.cancel_reason: update_data["cancel_reason"] = data.cancel_reason
    if data.status == "completado": update_data["completed_at"] = datetime.now(timezone.utc).isoformat()
    if data.vehicle_id: update_data["vehicle_id"] = data.vehicle_id

    if data.status == "en_curso" and data.mileage is not None: 
        if data.vehicle_id:
            vehicle = await db.vehicles.find_one({"id": data.vehicle_id})
            if vehicle and data.mileage < vehicle.get("mileage", 0):
                raise HTTPException(status_code=400, detail=f"Error Crítico: El kilometraje inicial ({data.mileage} km) NO puede ser menor al actual registrado ({vehicle.get('mileage', 0)} km).")
            await db.vehicles.update_one({"id": data.vehicle_id}, {"$set": {"mileage": data.mileage}})
        update_data["start_mileage"] = data.mileage
        
    if data.status == "completado" and data.mileage is not None: 
        start_km = trip.get("start_mileage", 0)
        if data.mileage <= start_km:
            raise HTTPException(status_code=400, detail=f"Error Crítico: El kilometraje final ({data.mileage} km) debe ser mayor al inicial ({start_km} km).")
        recorrido = data.mileage - start_km
        if recorrido > 1400:
            raise HTTPException(status_code=400, detail=f"Alerta Anti-Error: Es imposible recorrer {recorrido} km en un solo viaje.")
        update_data["end_mileage"] = data.mileage
        veh_id = trip.get("vehicle_id") or data.vehicle_id
        if veh_id: await db.vehicles.update_one({"id": veh_id}, {"$set": {"mileage": data.mileage}})

    await db.trips.update_one({"id": trip_id}, {"$set": update_data})
    
    # Manejo automático de estado de vehículo
    veh_id = trip.get("vehicle_id") or data.vehicle_id
    if veh_id:
        if data.status == "en_curso":
            await db.vehicles.update_one({"id": veh_id}, {"$set": {"status": "en_uso"}})
        elif data.status == "completado" or data.status == "cancelado":
            await db.vehicles.update_one({"id": veh_id}, {"$set": {"status": "disponible"}})
            
    # REGISTRO DE AUDITORÍA DE ESTADO
    detail_msg = f"Cambio de estado a {data.status} para folio {trip['tracking_number']}"
    if data.status == "cancelado" and data.cancel_reason: detail_msg += f". Motivo: {data.cancel_reason}"
    if data.mileage is not None: detail_msg += f". Km: {data.mileage}"
    
    await log_action(user["id"], user["name"], user["role"], data.status, "traslado", trip_id, detail_msg)
            
    return {"message": "Ok"}

@api_router.put("/trips/{trip_id}/group")
async def group_trip(trip_id: str, data: TripGroupUpdate, user=Depends(require_roles("conductor"))):
    trip = await db.trips.find_one({"id": trip_id})
    await db.trips.update_one({"id": trip_id}, {"$set": {"group_id": data.group_id, "order_in_group": data.order_in_group}})
    await log_action(user["id"], user["name"], user["role"], "agrupar", "traslado", trip_id, f"Traslado folio {trip.get('tracking_number')} agrupado en {data.group_id}")
    return {"message": "Ok"}

@api_router.delete("/trips/clear-all")
async def clear_all_trips(user=Depends(require_roles("admin"))):
    await db.trips.delete_many({})
    await log_action(user["id"], user["name"], user["role"], "limpiar_todo", "traslado", "all", "Se eliminaron todos los viajes de la base de datos")
    return {"message": "Todos los viajes han sido eliminados correctamente"}

@api_router.delete("/trips/{trip_id}")
async def delete_trip(trip_id: str, user=Depends(require_roles("admin"))):
    trip = await db.trips.find_one({"id": trip_id})

    if not trip: raise HTTPException(status_code=404, detail="Viaje no encontrado")
    
    await db.trips.delete_one({"id": trip_id})
    await log_action(user["id"], user["name"], user["role"], "eliminar", "traslado", trip_id, f"Eliminado viaje con folio {trip.get('tracking_number')}")
    return {"message": "Viaje eliminado correctamente"}


# ============ ENDPOINTS BITÁCORA ============

@api_router.post("/logbook/checklist")
async def create_checklist(data: VehicleChecklist, user=Depends(require_roles("conductor"))):
    doc = data.model_dump()
    doc["id"] = str(uuid.uuid4())
    doc["driver_id"] = user["id"]
    doc["driver_name"] = user["name"]
    doc["timestamp"] = datetime.now(timezone.utc).isoformat()
    doc["type"] = "checklist"
    await db.logbook.insert_one(doc)
    doc.pop("_id", None)
    return doc

@api_router.post("/logbook/fuel")
async def create_fuel_record(data: FuelRecord, user=Depends(require_roles("conductor"))):
    doc = data.model_dump()
    doc["id"] = str(uuid.uuid4())
    doc["driver_id"] = user["id"]
    doc["driver_name"] = user["name"]
    doc["timestamp"] = datetime.now(timezone.utc).isoformat()
    doc["type"] = "fuel"
    
    vehicle = await db.vehicles.find_one({"id": data.vehicle_id})
    if vehicle and data.mileage < vehicle.get("mileage", 0):
        raise HTTPException(status_code=400, detail=f"Error Crítico: Kilometraje ingresado ({data.mileage}) es menor al de la ambulancia ({vehicle.get('mileage', 0)}).")
    
    await db.vehicles.update_one({"id": data.vehicle_id}, {"$set": {"mileage": data.mileage}})
    await db.logbook.insert_one(doc)
    doc.pop("_id", None)
    return doc

@api_router.post("/logbook/incident")
async def create_incident(data: IncidentRecord, user=Depends(require_roles("conductor"))):
    doc = data.model_dump()
    doc["id"] = str(uuid.uuid4())
    doc["driver_id"] = user["id"]
    doc["driver_name"] = user["name"]
    doc["timestamp"] = datetime.now(timezone.utc).isoformat()
    doc["type"] = "incident"
    await db.logbook.insert_one(doc)
    doc.pop("_id", None)
    return doc

@api_router.get("/logbook/{vehicle_id}")
async def get_vehicle_logbook(vehicle_id: str, date: str = None, user=Depends(get_current_user)):
    target_date = date or datetime.now(timezone.utc).strftime("%Y-%m-%d")
    trips = await db.trips.find({
        "vehicle_id": vehicle_id, 
        "scheduled_date": target_date,
        "status": {"$ne": "cancelado"}
    }, {"_id": 0}).to_list(100)
    
    logs = await db.logbook.find({
        "vehicle_id": vehicle_id,
        "timestamp": {"$regex": f"^{target_date}"}
    }, {"_id": 0}).to_list(100)
    
    return {"trips": trips, "logs": logs}

# ============ OTHER ENDPOINTS ============

@api_router.get("/destinations")
async def list_destinations(user=Depends(get_current_user)):
    return await db.destinations.find({}, {"_id": 0}).to_list(1000)

@api_router.get("/stats")
async def get_stats(user=Depends(require_roles("admin", "coordinador", "gestion_camas"))):
    return {
        "total_trips": await db.trips.count_documents({}),
        "pending_trips": await db.trips.count_documents({"status": "pendiente"}),
        "active_trips": await db.trips.count_documents({"status": {"$in": ["asignado", "en_curso"]}}),
        "completed_trips": await db.trips.count_documents({"status": "completado"}),
        "total_vehicles": await db.vehicles.count_documents({}),
        "vehicles_available": await db.vehicles.count_documents({"status": "disponible"}),
        "vehicles_en_uso": await db.vehicles.count_documents({"status": "en_uso"}),
        "vehicles_fuera_de_servicio": await db.vehicles.count_documents({"status": "fuera_de_servicio"}),
        "total_drivers": await db.users.count_documents({"role": "conductor", "status": "aprobado"}),
        "pending_users": await db.users.count_documents({"status": "pendiente"})
    }

@api_router.get("/stats/advanced")
async def get_advanced_stats(user=Depends(require_roles("admin"))):
    now = datetime.now(timezone.utc)
    
    # --- Trends: trips per day for the last 30 days ---
    daily_trends = []
    for i in range(29, -1, -1):
        day = (now - timedelta(days=i)).strftime("%Y-%m-%d")
        count = await db.trips.count_documents({"scheduled_date": day})
        completed = await db.trips.count_documents({"scheduled_date": day, "status": "completado"})
        daily_trends.append({"date": day, "total": count, "completados": completed})
    
    # --- Distribution by status ---
    status_dist = []
    for s in ["pendiente", "por_visar", "asignado", "en_curso", "completado", "cancelado"]:
        c = await db.trips.count_documents({"status": s})
        if c > 0:
            status_dist.append({"name": s.replace("_", " ").title(), "value": c})
    
    # --- Distribution by trip type ---
    type_dist = []
    for t in ["clinico", "no_clinico"]:
        c = await db.trips.count_documents({"trip_type": t})
        if c > 0:
            label = "Clínico" if t == "clinico" else "No Clínico"
            type_dist.append({"name": label, "value": c})
    
    # --- Top 5 destinations ---
    all_trips = await db.trips.find({}, {"destination": 1, "_id": 0}).to_list(50000)
    dest_count = {}
    for t in all_trips:
        d = t.get("destination", "Desconocido")
        dest_count[d] = dest_count.get(d, 0) + 1
    top_destinations = sorted([{"name": k, "viajes": v} for k, v in dest_count.items()], key=lambda x: x["viajes"], reverse=True)[:5]
    
    # --- Top 5 drivers by completed trips ---
    completed_trips = await db.trips.find({"status": "completado", "driver_name": {"$ne": None}}, {"driver_name": 1, "_id": 0}).to_list(50000)
    driver_count = {}
    for t in completed_trips:
        d = t.get("driver_name", "Sin Conductor")
        driver_count[d] = driver_count.get(d, 0) + 1
    top_drivers = sorted([{"name": k, "viajes": v} for k, v in driver_count.items()], key=lambda x: x["viajes"], reverse=True)[:5]
    
    # --- Total mileage ---
    mileage_trips = await db.trips.find({"start_mileage": {"$exists": True}, "end_mileage": {"$exists": True}}, {"start_mileage": 1, "end_mileage": 1, "_id": 0}).to_list(50000)
    total_km = sum((t.get("end_mileage", 0) - t.get("start_mileage", 0)) for t in mileage_trips if t.get("end_mileage", 0) > t.get("start_mileage", 0))
    
    # --- Priority distribution ---
    priority_dist = []
    for p in ["urgente", "normal", "programado"]:
        c = await db.trips.count_documents({"priority": p})
        if c > 0:
            priority_dist.append({"name": p.title(), "value": c})
    
    # --- Cancelled trips count + rate ---
    cancelled = await db.trips.count_documents({"status": "cancelado"})
    total = await db.trips.count_documents({})
    cancel_rate = round((cancelled / total * 100), 1) if total > 0 else 0
    
    # --- Today's stats ---
    today = now.strftime("%Y-%m-%d")
    trips_today = await db.trips.count_documents({"scheduled_date": today})
    completed_today = await db.trips.count_documents({"scheduled_date": today, "status": "completado"})
    
    # --- Total users by role ---
    users_by_role = []
    for r in ["solicitante", "conductor", "coordinador", "gestion_camas", "admin"]:
        c = await db.users.count_documents({"role": r, "status": "aprobado"})
        label_map = {"solicitante": "Solicitantes", "conductor": "Conductores", "coordinador": "Coordinadores", "gestion_camas": "Gestión Camas", "admin": "Admins"}
        if c > 0:
            users_by_role.append({"name": label_map.get(r, r), "value": c})
    
    return {
        "daily_trends": daily_trends,
        "status_distribution": status_dist,
        "type_distribution": type_dist,
        "priority_distribution": priority_dist,
        "top_destinations": top_destinations,
        "top_drivers": top_drivers,
        "total_km": round(total_km, 1),
        "cancel_rate": cancel_rate,
        "cancelled_trips": cancelled,
        "trips_today": trips_today,
        "completed_today": completed_today,
        "users_by_role": users_by_role,
    }


@api_router.get("/audit-logs")
async def get_audit_logs(user=Depends(require_roles("admin"))):
    return await db.audit_logs.find({}, {"_id": 0}).sort("timestamp", -1).to_list(5000)

@api_router.post("/seed-admin")
async def seed_admin():
    if not await db.users.find_one({"role": "admin"}):
        await db.users.insert_one({"id": str(uuid.uuid4()), "email": "admin@hospital.cl", "password_hash": pwd_context.hash("admin123"), "name": "Administrador", "role": "admin", "status": "aprobado", "created_at": datetime.now(timezone.utc).isoformat()})
    return {"message": "Ok"}

@api_router.get("/clinical-staff")
async def list_clinical_staff(user=Depends(get_current_user)):
    return await db.clinical_staff.find({}, {"_id": 0}).to_list(1000)

@api_router.post("/clinical-staff")
async def create_clinical_staff(data: ClinicalStaffCreate, user=Depends(require_roles("admin", "gestion_camas"))):
    staff = data.model_dump()
    staff["id"] = str(uuid.uuid4())
    staff["created_at"] = datetime.now(timezone.utc).isoformat()
    await db.clinical_staff.insert_one(staff)
    staff.pop("_id", None)
    return staff

@api_router.put("/clinical-staff/{staff_id}")
async def update_clinical_staff(staff_id: str, data: ClinicalStaffUpdate, user=Depends(require_roles("admin", "gestion_camas"))):
    update_data = {k: v for k, v in data.model_dump().items() if v is not None}
    if not update_data:
        return {"message": "Sin cambios"}
    await db.clinical_staff.update_one({"id": staff_id}, {"$set": update_data})
    return {"message": "Personal clínico actualizado"}

@api_router.delete("/clinical-staff/{staff_id}")
async def delete_clinical_staff(staff_id: str, user=Depends(require_roles("admin", "gestion_camas"))):
    await db.clinical_staff.delete_one({"id": staff_id})
    return {"message": "Personal clínico eliminado"}

# ============ ORIGIN SERVICES CRUD ============

@api_router.get("/origin-services")
async def list_origin_services(user=Depends(get_current_user)):
    return await db.origin_services.find({}, {"_id": 0}).to_list(1000)

@api_router.post("/origin-services")
async def create_origin_service(data: OriginServiceCreate, user=Depends(require_roles("admin", "gestion_camas"))):
    service = data.model_dump()
    service["id"] = str(uuid.uuid4())
    service["created_at"] = datetime.now(timezone.utc).isoformat()
    await db.origin_services.insert_one(service)
    service.pop("_id", None)
    return service

@api_router.put("/origin-services/{service_id}")
async def update_origin_service(service_id: str, data: OriginServiceUpdate, user=Depends(require_roles("admin", "gestion_camas"))):
    update_data = {k: v for k, v in data.model_dump().items() if v is not None}
    if not update_data:
        return {"message": "Sin cambios"}
    await db.origin_services.update_one({"id": service_id}, {"$set": update_data})
    return {"message": "Servicio actualizado"}

@api_router.delete("/origin-services/{service_id}")
async def delete_origin_service(service_id: str, user=Depends(require_roles("admin", "gestion_camas"))):
    await db.origin_services.delete_one({"id": service_id})
    return {"message": "Servicio eliminado"}

# ============ RUT VALIDATION ============

@api_router.get("/validate-rut/{rut}")
async def validate_rut(rut: str):
    """Validate Chilean RUT using modulo 11 algorithm"""
    clean = rut.replace(".", "").replace("-", "").upper().strip()
    if len(clean) < 2:
        return {"valid": False, "formatted": rut}
    body = clean[:-1]
    dv = clean[-1]
    if not body.isdigit():
        return {"valid": False, "formatted": rut}
    total = 0
    factor = 2
    for digit in reversed(body):
        total += int(digit) * factor
        factor = factor + 1 if factor < 7 else 2
    remainder = 11 - (total % 11)
    if remainder == 11:
        expected = "0"
    elif remainder == 10:
        expected = "K"
    else:
        expected = str(remainder)
    valid = dv == expected
    # Format: XX.XXX.XXX-X
    formatted_body = ""
    for i, ch in enumerate(reversed(body)):
        if i > 0 and i % 3 == 0:
            formatted_body = "." + formatted_body
        formatted_body = ch + formatted_body
    formatted = f"{formatted_body}-{expected}"
    return {"valid": valid, "formatted": formatted}

# ============ DRIVER HISTORY ============

@api_router.get("/trips/v2/history")
async def driver_history_v2(user=Depends(get_current_user)):
    user_id = user.get("id")
    user_name = user.get("name")
    logger.info(f"!!! HISTORY_V2_START !!! user_id={user_id} name={user_name}")
    
    if not user_id:
        logger.error("DRIVER_HISTORY_ERROR: No user_id found in user object")
        return []

    query = {
        "$or": [
            {"driver_id": user_id},
            {"driver_name": user_name},
            {"previous_driver_id": user_id}
        ]
    }
    
    try:
        all_related = await db.trips.find(query, {"_id": 0}).to_list(2000)
        logger.info(f"DRIVER_HISTORY_DB: Found {len(all_related)} trips in DB")
    except Exception as e:
        logger.error(f"DRIVER_HISTORY_DB_ERROR: {str(e)}")
        return []
    
    history = []
    for t in all_related:
        stat = t.get("status")
        # Caso 1: Devuelto
        if t.get("previous_driver_id") == user_id:
            t["_history_status"] = "devuelto"
            history.append(t)
        # Caso 2: Finalizado / Completado
        elif stat in ["completado", "cancelado", "finalizado", "devuelto", "terminado"]:
            history.append(t)
        # Debug: veamos qué otros estados hay asignados a este driver
        else:
            logger.info(f"DRIVER_HISTORY_SKIP: trip={t.get('id')} status={stat}")

    def get_sort_key(x):
        return x.get("completed_at") or x.get("returned_at") or x.get("updated_at") or x.get("created_at") or ""

    history.sort(key=get_sort_key, reverse=True)
    logger.info(f"HISTORY_V2_FINAL: Returning {len(history)} items")
    
    return {"trips": history, "debug_user": user_id, "timestamp": datetime.now(timezone.utc).isoformat()}

# ============ DEBUG ENDPOINT (TEMPORAL) ============
@api_router.get("/debug/trips-status")
async def debug_trips_status():
    total = await db.trips.count_documents({})
    pendientes = await db.trips.count_documents({"status": "pendiente"})
    pendientes_sin_driver = await db.trips.count_documents({"status": "pendiente", "driver_id": None})
    pendientes_sin_driver_field = await db.trips.count_documents({"status": "pendiente", "driver_id": {"$exists": False}})
    asignados = await db.trips.count_documents({"status": "asignado"})
    en_curso = await db.trips.count_documents({"status": "en_curso"})
    completados = await db.trips.count_documents({"status": "completado"})
    
    # Sample de viajes pendientes
    sample_pendientes = []
    async for t in db.trips.find({"status": "pendiente"}, {"_id": 0, "id": 1, "tracking_number": 1, "driver_id": 1, "driver_name": 1, "status": 1, "origin": 1, "destination": 1}).limit(5):
        sample_pendientes.append(t)
    
    # Sample de viajes asignados
    sample_asignados = []
    async for t in db.trips.find({"status": "asignado"}, {"_id": 0, "id": 1, "tracking_number": 1, "driver_id": 1, "driver_name": 1, "status": 1}).limit(5):
        sample_asignados.append(t)
    
    # Conductores
    conductores = []
    async for u in db.users.find({"role": "conductor"}, {"_id": 0, "id": 1, "name": 1, "email": 1}).limit(10):
        conductores.append(u)

    return {
        "total_trips": total,
        "pendientes": pendientes,
        "pendientes_driver_null": pendientes_sin_driver,
        "pendientes_driver_no_field": pendientes_sin_driver_field,
        "asignados": asignados,
        "en_curso": en_curso,
        "completados": completados,
        "sample_pendientes": sample_pendientes,
        "sample_asignados": sample_asignados,
        "conductores": conductores
    }

# ============ COORDINATOR DASHBOARD STATS ============

@api_router.get("/stats/dashboard")
async def dashboard_stats(user=Depends(require_roles("coordinador", "admin", "gestion_camas"))):
    now = datetime.now(timezone.utc)
    today = now.strftime("%Y-%m-%d")
    # Start of week (Monday)
    start_of_week = (now - timedelta(days=now.weekday())).strftime("%Y-%m-%d")
    # Start of month
    start_of_month = now.strftime("%Y-%m-01")

    completed_today = await db.trips.count_documents({"status": "completado", "scheduled_date": today})
    completed_week = await db.trips.count_documents({"status": "completado", "scheduled_date": {"$gte": start_of_week, "$lte": today}})
    completed_month = await db.trips.count_documents({"status": "completado", "scheduled_date": {"$gte": start_of_month, "$lte": today}})

    total_all = await db.trips.count_documents({})
    by_status = {
        "pendiente": await db.trips.count_documents({"status": "pendiente"}),
        "revision_gestor": await db.trips.count_documents({"status": "revision_gestor"}),
        "asignado": await db.trips.count_documents({"status": "asignado"}),
        "en_curso": await db.trips.count_documents({"status": "en_curso"}),
        "completado": await db.trips.count_documents({"status": "completado"}),
        "cancelado": await db.trips.count_documents({"status": "cancelado"}),
    }

    active_drivers = await db.users.count_documents({"role": "conductor", "status": "aprobado"})
    # Drivers currently on a trip
    busy_driver_ids = await db.trips.distinct("driver_id", {"status": "en_curso"})
    busy_drivers = len([d for d in busy_driver_ids if d])

    # Average trip duration (from trips that have both start and completed_at)
    completed_trips = await db.trips.find(
        {"status": "completado", "start_mileage": {"$ne": None}, "end_mileage": {"$ne": None}},
        {"_id": 0, "start_mileage": 1, "end_mileage": 1}
    ).to_list(500)
    avg_km = 0
    if completed_trips:
        distances = [t["end_mileage"] - t["start_mileage"] for t in completed_trips if t.get("end_mileage") and t.get("start_mileage")]
        avg_km = round(sum(distances) / len(distances), 1) if distances else 0

    return {
        "completed_today": completed_today,
        "completed_week": completed_week,
        "completed_month": completed_month,
        "total_all": total_all,
        "by_status": by_status,
        "active_drivers": active_drivers,
        "busy_drivers": busy_drivers,
        "avg_km_per_trip": avg_km,
    }



# ============ REPORTS: LIBRO DE CONTROL DE RECORRIDO ============

import traceback
from html import escape

async def _fetch_logbook_data(vehicle_id: str, start_date: str, end_date: str) -> dict:
    vehicle = await db.vehicles.find_one({"id": vehicle_id}, {"_id": 0})
    if not vehicle:
        raise HTTPException(status_code=404, detail="Vehículo no encontrado")

    trips = await db.trips.find({
        "vehicle_id": vehicle_id,
        "scheduled_date": {"$gte": start_date, "$lte": end_date},
        "status": {"$in": ["completado", "en_curso", "asignado"]}
    }, {"_id": 0}).sort("scheduled_date", 1).to_list(1000)

    fuel = await db.logbook.find({
        "vehicle_id": vehicle_id,
        "type": "fuel",
        "timestamp": {"$gte": start_date, "$lte": f"{end_date}T23:59:59"}
    }, {"_id": 0}).sort("timestamp", 1).to_list(500)

    incidents = await db.logbook.find({
        "vehicle_id": vehicle_id,
        "type": "incident",
        "timestamp": {"$gte": start_date, "$lte": f"{end_date}T23:59:59"}
    }, {"_id": 0}).sort("timestamp", 1).to_list(500)

    for t in trips:
        try:
            audit = await db.audit_logs.find_one(
                {"entity_id": t["id"], "action": {"$in": ["aprobar", "manager_assign", "despachar"]}},
                {"_id": 0, "user_name": 1}
            )
            t["authorized_by"] = audit["user_name"] if audit else t.get("requester_name", "Sistema")
        except:
            t["authorized_by"] = t.get("requester_name", "Sistema")

    return {
        "vehicle": vehicle,
        "period": {"start": start_date, "end": end_date},
        "trips": trips,
        "fuel_logs": fuel,
        "incident_logs": incidents
    }

@api_router.get("/reports/logbook")
async def get_logbook_preview(vehicle_id: str, start_date: str, end_date: str, user=Depends(require_roles("admin", "coordinador"))):
    return await _fetch_logbook_data(vehicle_id, start_date, end_date)

@api_router.get("/reports/logbook-excel")
async def export_logbook_excel(vehicle_id: str, start_date: str, end_date: str, user=Depends(require_roles("admin", "coordinador"))):
    try:
        data = await _fetch_logbook_data(vehicle_id, start_date, end_date)
        v = data["vehicle"]
        wb = Workbook(); ws = wb.active; ws.title = "Libro de Recorrido"
        
        h_fill = PatternFill(start_color="1a5276", end_color="1a5276", fill_type="solid")
        white_font = Font(color="FFFFFF", bold=True)
        thin = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

        ws.merge_cells("A1:M1")
        ws["A1"] = f"LIBRO DE RECORRIDO - {v.get('plate', 'N/A')} ({v.get('brand', '')} {v.get('model', '')})"
        ws["A1"].font = Font(size=14, bold=True)
        ws["A1"].alignment = Alignment(horizontal="center")

        headers = ["Fecha", "Salida", "Llegada", "Km Ini", "Km Fin", "Km Rec", "Origen", "Destino", "Motivo", "Conductor", "Pasajeros", "Autorizado", "Folio"]
        for i, h in enumerate(headers, 1):
            c = ws.cell(row=3, column=i, value=h)
            c.fill = h_fill; c.font = white_font; c.border = thin

        row = 4
        total_km = 0.0
        for t in data["trips"]:
            try:
                s_km = float(t.get("start_mileage") or 0)
                e_km = float(t.get("end_mileage") or 0)
            except (ValueError, TypeError):
                s_km, e_km = 0.0, 0.0
                
            k_r = round(max(0, e_km - s_km), 1)
            total_km += k_r
            h_s = t.get("departure_time") or (t.get("created_at", "")[11:16] if "T" in t.get("created_at", "") else "")
            h_l = (t.get("completed_at", "")[11:16] if t.get("completed_at") and "T" in t.get("completed_at", "") else "")
            
            vals = [t.get("scheduled_date", ""), h_s, h_l, s_km, e_km, k_r, t.get("origin", ""), t.get("destination", ""), t.get("transfer_reason", ""), t.get("driver_name", ""), t.get("clinical_team", ""), t.get("authorized_by", ""), t.get("tracking_number", "")]
            for i, val in enumerate(vals, 1):
                cell = ws.cell(row=row, column=i, value=val)
                cell.border = thin
            row += 1
        
        ws.cell(row=row, column=5, value="TOTAL KM:").font = Font(bold=True)
        ws.cell(row=row, column=6, value=total_km).font = Font(bold=True)

        output = io.BytesIO(); wb.save(output); content = output.getvalue(); output.close()
        return Response(content=content, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f"attachment; filename=Libro_{v.get('plate', 'VEH')}_{start_date}.xlsx"})
    except Exception as e:
        logger.error(f"Excel error: {e}")
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))

@api_router.get("/reports/logbook-pdf")
async def export_logbook_pdf(vehicle_id: str, start_date: str, end_date: str, user=Depends(require_roles("admin", "coordinador"))):
    try:
        data = await _fetch_logbook_data(vehicle_id, start_date, end_date)
        v = data["vehicle"]
        output = io.BytesIO()
        doc = SimpleDocTemplate(output, pagesize=landscape(letter), topMargin=10*mm, bottomMargin=10*mm)
        styles = getSampleStyleSheet()
        styles.add(ParagraphStyle(name="CT", fontSize=6.5, leading=8))
        styles.add(ParagraphStyle(name="CB", fontSize=8, fontName="Helvetica-Bold"))
        
        elements = [Paragraph(f"LIBRO DE RECORRIDO - HOSPITAL DE CURICÓ", styles["Title"]), Spacer(1, 5*mm)]
        elements.append(Paragraph(f"Vehículo: {escape(v.get('plate', 'N/A'))} ({escape(v.get('brand', ''))} {escape(v.get('model', ''))})&nbsp;&nbsp;&nbsp;Período: {start_date} al {end_date}", styles["Normal"]))
        elements.append(Spacer(1, 5*mm))

        h = ["Fecha", "Salida", "Llegada", "Km Ini", "Km Fin", "Km Rec", "Origen", "Destino", "Motivo", "Conductor", "Autorizado"]
        t_data = [[Paragraph(x, styles["CB"]) for x in h]]
        total_km = 0.0
        for t in data["trips"]:
            try:
                s_km = float(t.get("start_mileage") or 0)
                e_km = float(t.get("end_mileage") or 0)
            except (ValueError, TypeError):
                s_km, e_km = 0.0, 0.0
            k_r = round(max(0, e_km - s_km), 1)
            total_km += k_r
            
            orig = escape((t.get("origin") or "")[:25])
            dest = escape((t.get("destination") or "")[:25])
            motv = escape((t.get("transfer_reason") or t.get("task_details") or "")[:35])
            drvr = escape((t.get("driver_name") or "")[:20])
            auth = escape((t.get("authorized_by") or "")[:20])
            
            t_data.append([
                t.get("scheduled_date", ""), "", "", str(s_km), str(e_km), str(k_r), 
                Paragraph(orig, styles["CT"]), Paragraph(dest, styles["CT"]), 
                Paragraph(motv, styles["CT"]), Paragraph(drvr, styles["CT"]), 
                Paragraph(auth, styles["CT"])
            ])

        t = Table(t_data, colWidths=[55, 35, 35, 40, 40, 40, 75, 75, 90, 70, 70])
        t.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1a5276")), ("TEXTCOLOR", (0, 0), (-1, 0), colors.white), ("GRID", (0, 0), (-1, -1), 0.5, colors.grey), ("FONTSIZE", (0, 0), (-1, -1), 6.5), ("VALIGN", (0, 0), (-1, -1), "MIDDLE")]))
        elements.append(t)
        
        if data["incident_logs"]:
            elements.append(Spacer(1, 10*mm))
            elements.append(Paragraph("OBSERVACIONES / NOVEDADES", styles["CB"]))
            for inc in data["incident_logs"]:
                elements.append(Paragraph(f"- {str(inc.get('timestamp',''))[:10]}: {escape(str(inc.get('description','')))}", styles["CT"]))

        elements.append(Spacer(1, 15*mm))
        elements.append(Paragraph("_______________________________&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;_______________________________", styles["Normal"]))
        elements.append(Paragraph("<b>FIRMA CONDUCTOR&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;FIRMA JEFATURA</b>", styles["Normal"]))

        doc.build(elements); content = output.getvalue(); output.close()
        return Response(content=content, media_type="application/pdf", headers={"Content-Disposition": f"attachment; filename=Libro_{v.get('plate', 'VEH')}_{start_date}.pdf"})
    except Exception as e:
        logger.error(f"PDF error: {e}")
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))

app.include_router(api_router)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
